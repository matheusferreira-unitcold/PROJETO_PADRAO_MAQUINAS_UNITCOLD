from itertools import product
from openpyxl import load_workbook

def gerar_combinacoes_AC_QPC():
    # Configurações do arquivo
    caminho_arquivo = "C:\\Users\\unitc\\Unitcold\\TI Unitcold - Documentos\\PROJETO_PADRAO_MAQUINAS_UNITCOLD\\Gerador_de_Modulos\\Elaboracao_de_descricoes_de_modulos.xlsx"
    nome_aba = "AC-QPC"
    coluna_destino = 1  # Coluna A
    
    try:
        # Carrega a planilha existente
        planilha = load_workbook(caminho_arquivo)
        aba = planilha[nome_aba]
        
        # Definindo todas as opções por coluna
        configuracoes = {
            'familias': ["AC-QPC"],
            'medida': ["BxHxP"],
            'lado_conexao': ["0", "B", "U", "D", "L", "R"],
            'abertura_ihm': ["0", "PGE", "PGX", "E"],
            'chave_seletora_3p': ["0", "1"],
            'config_ab_22mm': ["0", "CxR"],
            'exaustor': ["0", "R", "L","F"],
            'veneziana': ["0","R","L","F"],
            'finalizacoes': ["X"]
        }
        
        # Calcula o total estimado de combinações (para progresso)
        total_combinacoes = (
            len(configuracoes['familias']) *
            len(configuracoes['medida']) *
            len(configuracoes['lado_conexao']) *
            len(configuracoes['abertura_ihm']) *
            len(configuracoes['chave_seletora_3p']) *
            len(configuracoes['config_ab_22mm']) *
            len(configuracoes['exaustor']) *
            len(configuracoes ['veneziana'])*
            len(configuracoes['finalizacoes'])
        )
        print(f"Gerando {total_combinacoes:,} combinações possíveis...")

        # Gerando todas as combinações
        todas_combinacoes = product(
            configuracoes['familias'],
            configuracoes['medida'],
            configuracoes['lado_conexao'],
            configuracoes['abertura_ihm'],
            configuracoes['chave_seletora_3p'],
            configuracoes['config_ab_22mm'],
            configuracoes['exaustor'],
            configuracoes['veneziana'],
            configuracoes['finalizacoes']
        )

        # Encontrando a última linha preenchida
        linha_inicio = aba.max_row + 1
        
        # Adicionando os dados na planilha
        for i, combo in enumerate(todas_combinacoes, start=linha_inicio):
            # Filtra campos vazios e junta com "-"
            codigo = "-".join(valor for valor in combo if valor)
            aba.cell(row=i, column=coluna_destino, value=codigo)
            
            # Mostra progresso a cada 1000 registros
            if i % 1000 == 0:
                print(f"Processadas {i - linha_inicio + 1:,} combinações...")

        # Salva as alterações
        planilha.save(caminho_arquivo)
        print(f"\n✅ {aba.max_row - linha_inicio + 1:,} combinações adicionadas com sucesso na aba '{nome_aba}'!")
        
    except Exception as e:
        print(f"\n❌ Erro durante a execução: {str(e)}")
        if 'planilha' in locals():
            planilha.close()

if __name__ == "__main__":
    gerar_combinacoes_AC_QPC()