from itertools import product
from openpyxl import load_workbook

def gerar_combinacoes_CF_AHU_MT():
    # Configurações do arquivo
    caminho_arquivo = "C:\\Users\\unitc\\Unitcold\\TI Unitcold - Documentos\\PROJETO_PADRAO_MAQUINAS_UNITCOLD\\Gerador_de_Modulos\\Elaboracao_de_descricoes_de_modulos.xlsx"
    nome_aba = "CF-AHU-MT"
    coluna_destino = 1  # Coluna A
    
    try:
        # Carrega a planilha existente
        planilha = load_workbook(caminho_arquivo)
        aba = planilha[nome_aba]
        
        # Definindo todas as opções por coluna
        configuracoes = {
            'familias': ["CF-AHU-MT"],
            'gabinete': ["S","P"],
            'tam_gabinete': ["013", "020", "027", "034", "051", "068", "102", "136", "170", "204", "238", "272", "306", "340", "408"],
            'base_inferior': ["0", "F", "Q", "E"],
            'pos_gabinete': ["V", "H"],
            'acesso_manut': ["R", "L"],
            'conex_frigorifica': ["R", "L"],
            'conex_dreno': ["S", "D"],
            'trocador_calor': ["0", "DX/n", "2R", "4R", "6R", "8R", "An"],
            'disp_contr_refrig': ["0", "TV", "EV", "DK", "GE", "LG", "SS","TN","2V","3V","PICV"],
            'acessorios': ["AE", "UE", "FU","UV", "KC"],
            'finalizacoes': ["X"]
        }
        
        # Calcula o total estimado de combinações (para progresso)
        total_combinacoes = (
            len(configuracoes['familias']) *
            len(configuracoes['gabinete']) *
            len(configuracoes['tam_gabinete']) *
            len(configuracoes['base_inferior']) *
            len(configuracoes['pos_gabinete']) *
            len(configuracoes['acesso_manut']) *
            len(configuracoes['conex_frigorifica']) *
            len(configuracoes['conex_dreno']) *
            len(configuracoes['trocador_calor']) *
            len(configuracoes['disp_contr_refrig']) *
            len(configuracoes['acessorios']) *
            len(configuracoes['finalizacoes'])
        )
        print(f"Gerando {total_combinacoes:,} combinações possíveis...")

        # Gerando todas as combinações
        todas_combinacoes = product(
            configuracoes['familias'],
            configuracoes['gabinete'],
            configuracoes['tam_gabinete'],
            configuracoes['base_inferior'],
            configuracoes['pos_gabinete'],
            configuracoes['conex_frigorifica'],
            configuracoes['acesso_manut'],
            configuracoes['conex_dreno'],
            configuracoes['trocador_calor'],
            configuracoes['disp_contr_refrig'],
            configuracoes['acessorios'],
            configuracoes['finalizacoes']
        )

        # Encontrando a última linha preenchida
        linha_inicio = aba.max_row + 1
        
        # Adicionando os dados na planilha
        for i, combo in enumerate(todas_combinacoes, start=linha_inicio):
            # Filtra campos vazios e junta com "-"
            codigo = "-".join(valor for valor in combo if valor)
            aba.cell(row=i, column=coluna_destino, value=codigo)
            
            # Mostra progresso a cada 1000 registros
            if i % 1000 == 0:
                print(f"Processadas {i - linha_inicio + 1:,} combinações...")

        # Salva as alterações
        planilha.save(caminho_arquivo)
        print(f"\n✅ {aba.max_row - linha_inicio + 1:,} combinações adicionadas com sucesso na aba '{nome_aba}'!")
        
    except Exception as e:
        print(f"\n❌ Erro durante a execução: {str(e)}")
        if 'planilha' in locals():
            planilha.close()

if __name__ == "__main__":
    gerar_combinacoes_CF_AHU_MT()