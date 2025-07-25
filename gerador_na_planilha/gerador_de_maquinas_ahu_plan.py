from itertools import product
from openpyxl import load_workbook

def gerar_combinacoes_ahu():
    # Configurações do arquivo
    caminho_arquivo = "C:\\Users\\unitc\\Unitcold\\TI Unitcold - Documentos\\PROJETO_PADRAO_MAQUINAS_UNITCOLD\\Elaboracao_de_descricoes_de_maquinas.xlsx"
    nome_aba = "AHU"
    coluna_destino = 1  # Coluna A
    
    try:
        # Carrega a planilha existente
        planilha = load_workbook(caminho_arquivo)
        aba = planilha[nome_aba]
        
        # Definindo todas as opções por coluna
        configuracoes = {
            'familias': ["CR-AHU"],
            'modelos': ["006", "013", "020", "027", "034", "051", "068", 
                       "102", "136", "170", "204", "238", "272", "306", "340", "408"],
            'gabinete': ["S", "P", ""],
            'trocador': ["1,0/1", "4R"],
            'parceiros': ["SS", "TN", "DK", "LG"],
            'ventilador': ["", "LL", "PF", "EC"],
            'filtros': ["G4", "M5", "F8"],
            'finalizacoes': ["X"]
        }
        
        # Calcula o total estimado de combinações (para progresso)
        total_combinacoes = (
            len(configuracoes['familias']) *
            len(configuracoes['modelos']) *
            len(configuracoes['gabinete']) *
            len(configuracoes['trocador']) *
            len(configuracoes['parceiros']) *
            len(configuracoes['ventilador']) *
            len(configuracoes['filtros']) *
            len(configuracoes['finalizacoes'])
        )
        print(f"Gerando {total_combinacoes:,} combinações possíveis...")

        # Gerando todas as combinações
        todas_combinacoes = product(
            configuracoes['familias'],
            configuracoes['modelos'],
            configuracoes['gabinete'],
            configuracoes['trocador'],
            configuracoes['parceiros'],
            configuracoes['ventilador'],
            configuracoes['filtros'],
            configuracoes['finalizacoes']
        )

        # Encontrando a última linha preenchida
        linha_inicio = aba.max_row + 1
        
        # Adicionando os dados na planilha
        for i, combo in enumerate(todas_combinacoes, start=linha_inicio):
            # Filtra campos vazios e junta com "-"
            codigo = "-".join(valor for valor in combo if valor)
            aba.cell(row=i, column=coluna_destino, value=codigo)
            
            # Mostra progresso a cada 1000 registros
            if i % 1000 == 0:
                print(f"Processadas {i - linha_inicio + 1:,} combinações...")

        # Salva as alterações
        planilha.save(caminho_arquivo)
        print(f"\n✅ {aba.max_row - linha_inicio + 1:,} combinações adicionadas com sucesso na aba '{nome_aba}'!")
        
    except Exception as e:
        print(f"\n❌ Erro durante a execução: {str(e)}")
        if 'planilha' in locals():
            planilha.close()

if __name__ == "__main__":
    gerar_combinacoes_ahu()